from datetime import datetime
from time import time
import os, sys, platform
import threading
import openpyxl
from json_analyzer import compare_json

log = open('parse_config.log', 'w')

def TextFSM(raw_data, Hostname, Command, device_type):
	#Import textfsm library
	import textfsm

	# Open textfsm template
	if device_type == 'cisco_ios':
		if Command == 'show inventory':
			template_name = "./textfsm/ios_show_inventory.textfsm"
		elif Command == 'show version':
			template_name = "./textfsm/ios_show_version.textfsm"
		elif Command == 'show vlan':
			template_name = "./textfsm/ios_show_vlan.textfsm"
		elif Command == 'show cdp neighbor':
			template_name = "./textfsm/ios_show_cdp_neighbors.textfsm"
		elif Command == 'show cdp neighbor detail':
			template_name = "./textfsm/ios_show_cdp_neighbor_detail.textfsm"
		elif Command == 'show interface':
			template_name = "./textfsm/ios_show_interface.textfsm"
		elif Command == 'show ip ospf neighbor':
			template_name = "./textfsm/ios_show_ip_ospf_neighbor.textfsm"
		elif Command == 'show ip interface brief':
			template_name = "./textfsm/ios_show_ip_int_brief.textfsm"
		elif Command == 'show ip bgp summary':
			template_name = "./textfsm/ios_show_ip_bgp_summary.textfsm"
		elif Command == 'show interface status':
			template_name = "./textfsm/ios_show_interfaces_status.textfsm"
		elif Command == 'show mac-address-table':
			template_name = "./textfsm/ios_show_mac-address-table.textfsm"
		elif Command == 'show ip arp':
			template_name = "./textfsm/ios_show_ip_arp.textfsm"
		elif Command == 'show spanning-tree interface':
			template_name = "./textfsm/ios_show_spanning-tree_interface.textfsm"
		elif Command == 'show etherchannel summary':
			template_name = "./textfsm/ios_show_etherchannel_summary.textfsm"
		else:
			print >> log, ("Command %s is not matched with Template" % (Command))
			print ("Template Not Found")
			return

	elif device_type == 'cisco_nxos':
		if Command == 'show inventory':
			template_name = "./textfsm/nxos_show_inventory.textfsm"
		elif Command == 'show version':
			template_name = "./textfsm/nxos_show_version.textfsm"
		elif Command == 'show vlan':
			template_name = "./textfsm/nxos_show_vlan.textfsm"
		elif Command == 'show cdp neighbor':
			template_name = "./textfsm/nxos_show_cdp_neighbor.textfsm"
		elif Command == 'show interface':
			template_name = "./textfsm/nxos_show_interface.textfsm"
		elif Command == 'show fex':
			template_name = "./textfsm/nxos_show_fex.textfsm"
		elif Command == 'show fex detail':
			template_name = "./textfsm/nxos_show_fex_detail.textfsm"
		elif Command == 'show interface status':
			template_name = "./textfsm/nxos_show_interface_status.textfsm"
		elif Command == 'show port-channel summary':
			template_name = "./textfsm/nxos_show_port-channel_summary.textfsm"
		elif Command == 'show processes cpu':
			template_name = "./textfsm/nxos_show_processes_cpu.textfsm"
		elif Command == 'show feature':
			template_name = "./textfsm/nxos_show_feature.textfsm"
		elif Command == 'show license usage':
			template_name = "./textfsm/nxos_show_license_usage.textfsm"
		elif Command == 'show mac-address-table':
			template_name = "./textfsm/nxos_show_mac_address-table.textfsm"
		elif Command == 'show ip arp':
			template_name = "./textfsm/nxos_show_ip_arp.textfsm"
		else:
			print >> log, ("Command %s is not matched with Template" % (Command))
			print ("Template Not Found")
			return

	elif device_type == 'arista_eos':
		if Command == 'show version':
			template_name = "./textfsm/arista_eos_show_version.textfsm"
		elif Command == 'show interfaces':
			template_name = "./textfsm/arista_eos_show_interfaces.textfsm"
		elif Command == 'show interface status':
			template_name = "./textfsm/arista_eos_show_interfaces_status.textfsm"
		elif Command == 'show ip interface brief':
			template_name = "./textfsm/arista_eos_show_ip_interface_brief.textfsm"
		elif Command == 'show lldp neighbor':
			template_name = "./textfsm/arista_eos_show_lldp_neighbors.textfsm"
		elif Command == 'show mlag':
			template_name = "./textfsm/arista_eos_show_mlag.textfsm"
		elif Command == 'show vlan':
			template_name = "./textfsm/arista_eos_show_vlan.textfsm"
		elif Command == 'show boot-config':
			template_name = "./textfsm/arista_eos_show_boot-config.textfsm"
		elif Command == 'show hostname':
			template_name = "./textfsm/arista_eos_show_hostname.textfsm"
		elif Command == 'show lldp neighbor detail':
			template_name = "./textfsm/arista_eos_show_lldp_neighbors_detail.textfsm"
		elif Command == 'show inventory':
			template_name = "./textfsm/arista_eos_show_inventory.textfsm"
		else:
			print >> log, ("Command %s is not matched with Template" % (Command))
			print ("Template Not Found")
			return

	else:
		print >> log, ("Command %s is not matched with Template" % (Command))
		print ("Template Not Found")
		return

	template = open(template_name)
	re_table = textfsm.TextFSM(template)
	fsm_results = re_table.ParseText(raw_data)

	# Write to JSON and XLSX
	Write_JSON(Hostname, Command, re_table, fsm_results)
	Write_XLSX_File(Hostname, Command, re_table, fsm_results)

	#thread = []

	#t1 = threading.Thread(target=Write_JSON, args=(Hostname, Command, re_table, fsm_results))
	#thread.append(t1)
	#t1.start()

	#t2 = threading.Thread(target=Write_XLSX_File, args=(Hostname, Command, re_table, fsm_results))
	#thread.append(t2)
	#t2.start()

	print >> log, ("Parsing %s of %s is completed" % (Command, Hostname))
	print ("Parsing %s of %s is completed" % (Command, Hostname))

def Write_JSON(Hostname, Command, re_table, fsm_results):
	import json
	from collections import OrderedDict

	isAnalyzed = raw_input('Do you want to compare JSON output with existing baseline?(Yes/No) : ').lower()
	Baseline_Filename = ""

	# Check Path existence, if not exist, create a new directory
	path = "./output/%s" % Hostname
	if not os.path.exists(path):
		os.mkdir(path)

	Format = '%Y%m%d'
	Filename = './output/%s/%s-%s.json' % (Hostname, Hostname, Command)

	#Convert Headers into List
	Headers = []
	for Header in re_table.header:
		Headers.append(Header)

	# Initialize the Dictionary List
	JSON_Result = []

	# Extract values from FSM Result, and transpose into Dictionary
	for Values in fsm_results:
	    JSON_Result.append(OrderedDict(zip(Headers, Values)))

	# Write into JSON File
	#print json.dumps(JSON_Result, indent=4)
	with open(Filename, 'w') as JSON_Output:
		json.dump(JSON_Result, JSON_Output, indent=3)

	# Print Logs
	print >> log, ("Writing JSON Formatted file of %s - %s is completed" % (Hostname, Command))
	print ("Writing JSON Formatted file of %s - %s is completed" % (Hostname, Command))

	if (isAnalyzed == 'y' or isAnalyzed == 'yes'):
		compare_json(Hostname, Command, Baseline_Filename, JSON_Output)

def Write_Summary_Excel(Hostname):
	#Point to the directory
	Format = '%Y%m%d'
	Filename = './output/%s/%s_%s.xlsx' % (Hostname, Hostname, datetime.now().strftime(Format))

	#Create workbook and worksheet
	wb = openpyxl.load_workbook(Filename)
	ws = wb.create_sheet('Summary')

	# Create the Interface Name
	Interface_Sheet = wb.get_sheet_by_name('show interface status')
	Interfaces = []
	for row in range(2, Interface_Sheet.max_row):
		Interfaces.append(Interface_Sheet.cell(row=row, column=1).value)

	#Create Header table
	ws.cell(row=1, column=1).value = "Port"
	ws.cell(row=1, column=2).value = "Description"
	ws.cell(row=1, column=3).value = "Status"
	ws.cell(row=1, column=4).value = "VLAN"
	ws.cell(row=1, column=5).value = "Type"
	ws.cell(row=1, column=6).value = "Detected MAC Address"
	ws.cell(row=1, column=7).value = "Detected IP Address"

	#Create the Initial formula
	ws.cell(row=2, column=1).value = "Eth1/1"
	ws.cell(row=2, column=2).value = "=VLOOKUP(A2,'show interface status'!$A$1:$O$1000,2,FALSE)"
	ws.cell(row=2, column=3).value = "=VLOOKUP(A2,'show interface status'!$A$1:$G$1000,3,FALSE)"
	ws.cell(row=2, column=4).value = "=VLOOKUP(A2,'show interface status'!$A$1:$G$1000,4,FALSE)"
	ws.cell(row=2, column=5).value = "=VLOOKUP(A2,'show interface status'!$A$1:$G$1000,7,FALSE)"
	ws.cell(row=2, column=6).value = "=VLOOKUP(A2,'show mac-address-table'!$A$1:$F$5000,3,FALSE)"
	ws.cell(row=2, column=7).value = "=VLOOKUP(F2,'show ip arp'!$A$1:$E$5000,2,FALSE)"

	# Initialize worksheet
	Row = 2

	for Interface in Interfaces:
		ws.cell(row=Row, column=1).value = Interface
		Row += 1
		#print row
	#Adjust the column width automatically based on value length
	dims = {}
	for row in ws.rows:
		for cell in row:
			if cell.value:
				dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))

	for col, value in dims.items():
		ws.column_dimensions[col].width = value

	wb.save(filename=Filename)

	# Print Logs
	print >> log, ("Writing Summary page of %s is completed" % (Hostname))
	print ("Writing Summary page of %s is completed" % (Hostname))

def Write_XLSX_File(Hostname, Command, re_table, fsm_results):
	# Check Path existence, if not exist, create a new directory
	path = "./output/%s" % Hostname
	if not os.path.exists(path):
		os.mkdir(path)

	Format = '%Y%m%d'
	Filename = './output/%s/%s_%s.xlsx' % (Hostname, Hostname, datetime.now().strftime(Format))

	#Check XLSX file existence, if not exist create a new one
	if not os.path.exists(Filename):
		wb = openpyxl.Workbook()
		ws = wb.active
		ws.title = Command
	else:
		wb = openpyxl.load_workbook(Filename)
		ws = wb.create_sheet(Command)

	# Initialize worksheet
	Row = 1
	Col = 1

	# Create header table
	for FSMHeader in re_table.header:
		ws.cell(row=Row, column=Col).value = FSMHeader
		Col += 1
	Row += 1

	#print fsm_results
	# Create table
	for contents in fsm_results:
		Col = 1

		for content in contents:
			if type(content) is list:
				for data in content:
					ws.cell(row=Row, column=Col).value = data
					Col += 1
			else:
				ws.cell(row=Row, column=Col).value = content
				Col += 1
		Row += 1

	#Adjust the column width automatically based on value length
	dims = {}
	for row in ws.rows:
		for cell in row:
			if cell.value:
				dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))

	for col, value in dims.items():
		ws.column_dimensions[col].width = value

	#Saving the workbook
	wb.save(filename=Filename)

	# Print Success Message
	print >> log, ("Writing XLS Formatted file of %s - %s is completed" % (Hostname, Command))
	print ("Writing XLS Formatted file of %s - %s is completed" % (Hostname, Command))

def main():
	t0 = time()
	print ("Start Parsing Configuration...")

	# User intervention for parse to spreadsheet and write to docx selection
	isSummarised = raw_input('Do you want to summarise interfaces information?(Yes/No) : ').lower()

	with open("./File/file.csv", "r") as SourceText:
		thread = []
		Hostnames = set()

		#Writing TextFSM to Excel
		for line in SourceText:
			FileName, hostname, command, device_type, end = line.split(',')
			Hostnames.add(hostname)
			input_file = open(FileName)
			raw_text_data = input_file.read()
			input_file.close()

			TextFSM(raw_text_data,hostname,command,device_type)
			#t = threading.Thread(target=TextFSM, args=(raw_text_data, hostname, command, device_type))
			#thread.append(t)
			#t.start()

		#Writing Summary Page to Excel
		if (isSummarised == 'y' or isSummarised == 'yes'):
			for Hostname in Hostnames:
				Write_Summary_Excel(Hostname)
				#t = threading.Thread(target=Write_Summary_Excel, args=(Hostname))
				#thread.append(t)
				#t.start()

	print "Total Parsing time:", round(time()-t0, 3), "s"

if __name__ == "__main__":
	# If this Python file runs by itself, run below command. If imported, this section is not run
	main()
